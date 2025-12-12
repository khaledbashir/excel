#!/usr/bin/env python3
"""
Quick demo of the Excel Agent in action.
This script demonstrates a simple Excel task using the agent.
"""

import asyncio
import sys
from pathlib import Path

async def demo_excel_agent():
    """Demonstrate the Excel Agent with a simple task."""
    
    print("🚀 Excel Agent Demo")
    print("=" * 50)
    
    # Check if we have the required files
    if not Path("test_input.xlsx").exists():
        print("❌ test_input.xlsx not found. Please run the setup first.")
        return False
    
    try:
        from excel_agent.agent_runner import ExcelAgentRunner, TaskInput
        from excel_agent.config import ExperimentConfig
        print("✓ Successfully imported Excel Agent modules")
    except ImportError as e:
        print(f"❌ Import error: {e}")
        return False
    
    # Note: This demo shows the structure but won't run without API keys
    print("\n📝 Demo Task Configuration:")
    print("-" * 30)
    
    # Example configuration (would need API key to actually run)
    config = ExperimentConfig(model='openai/gpt-4')
    print(f"✓ Model: {config.model}")
    print(f"✓ Timeout: {config.timeout}s")
    
    # Example task
    task_input = TaskInput(
        instruction="Add a new row with data: Alice, 28, Paris",
        input_file="test_input.xlsx",
        output_file="test_output.xlsx",
    )
    
    print(f"✓ Task: {task_input.instruction}")
    print(f"✓ Input: {task_input.input_file}")
    print(f"✓ Output: {task_input.output_file}")
    
    # Show what the agent would do
    print("\n🔍 What the Agent Will Do:")
    print("-" * 30)
    print("1. Connect to MCP server at http://127.0.0.1:8765/sse")
    print("2. Load the Excel file: test_input.xlsx")
    print("3. Analyze the current data structure")
    print("4. Add new row with: Alice, 28, Paris")
    print("5. Save the result to: test_output.xlsx")
    
    print("\n💡 To Run This Demo:")
    print("-" * 30)
    print("1. Set your API key: export OPENAI_API_KEY='your-key'")
    print("2. Start MCP server: python -c \"import asyncio; from excel_mcp.excel_server import mcp; asyncio.run(mcp.run_sse_async(host='127.0.0.1', port=8765))\"")
    print("3. Run this script with API key configured")
    
    # Show current Excel data
    try:
        import openpyxl
        wb = openpyxl.load_workbook("test_input.xlsx")
        ws = wb.active
        
        print(f"\n📊 Current Excel Data:")
        print("-" * 30)
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                print(" | ".join(str(cell) if cell is not None else "" for cell in row))
                
    except Exception as e:
        print(f"⚠️  Could not read Excel file: {e}")
    
    return True

def show_available_tools():
    """Show some of the available Excel tools."""
    print("\n🛠️  Available Excel Tools:")
    print("-" * 30)
    
    tools = [
        "📖 read_range_data - Read data from cell ranges",
        "✏️  set_range_data - Write data to cell ranges", 
        "🔍 search_cells - Find specific values",
        "➕ insert_rows/columns - Add rows or columns",
        "❌ delete_rows/columns - Remove rows or columns",
        "🎨 set_range_style - Apply formatting",
        "📋 create_sheet - Add new worksheets",
        "🏷️  rename_sheet - Rename worksheets",
        "🔗 merge_cells - Combine cells",
        "📊 add_conditional_formatting - Add format rules",
        "✅ add_data_validation - Set input rules",
        "🔄 autofill_range - Auto-fill series",
    ]
    
    for tool in tools:
        print(f"  {tool}")

async def main():
    """Run the demo."""
    success = await demo_excel_agent()
    show_available_tools()
    
    if success:
        print(f"\n🎉 Demo setup complete! Check USAGE_GUIDE.md for next steps.")
    else:
        print(f"\n❌ Demo failed. Please check the errors above.")
    
    return success

if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)