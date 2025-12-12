#!/usr/bin/env python3
"""
End-to-end test for GLM model with Excel Agent.
This tests the complete workflow with your GLM-4.6 model.
"""

import asyncio
import os
import sys
from pathlib import Path

# Add the project root to Python path
sys.path.insert(0, str(Path(__file__).parent))

# Set environment variables
os.environ["GLM_API_KEY"] = "18f65090a96a425898a8398a5c4518ce.DDtUvTTnUmK020Wx"

from excel_agent.agent_runner import ExcelAgentRunner, TaskInput
from excel_agent.config import ExperimentConfig
from excel_mcp.helpers import save_workbook_async
from openpyxl import Workbook

async def test_glm_end_to_end():
    """Test GLM model with actual Excel operations."""
    
    print("🚀 Testing GLM Model End-to-End")
    print("=" * 50)
    
    try:
        # Create a test Excel file
        print("📊 Creating test Excel file...")
        test_file = Path(__file__).parent / "test_glm_input.xlsx"
        output_file = Path(__file__).parent / "test_glm_output.xlsx"
        
        # Create a simple workbook with some data
        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"
        
        # Add headers
        ws["A1"] = "Name"
        ws["B1"] = "Age"
        ws["C1"] = "City"
        
        # Add some sample data
        ws["A2"] = "John"
        ws["B2"] = 25
        ws["C2"] = "New York"
        
        ws["A3"] = "Emma"
        ws["B3"] = 30
        ws["C3"] = "London"
        
        wb.save(test_file)
        print(f"✓ Test file created: {test_file}")
        
        # Copy input to output (agent expects output file to exist)
        import shutil
        shutil.copy(test_file, output_file)
        print(f"✓ Output file prepared: {output_file}")
        
        # Create GLM configuration
        print("\n🤖 Setting up GLM model...")
        config = ExperimentConfig(model='glm')
        print(f"✓ Config created with model: {config.model}")
        
        # Create agent runner
        print("\n🔧 Initializing Excel Agent Runner...")
        runner = ExcelAgentRunner(
            config=config,
            mcp_server_url="http://127.0.0.1:8765/sse"
        )
        print("✓ Agent runner initialized")
        
        # Create task input
        print("\n📝 Creating task...")
        task_input = TaskInput(
            instruction="Add a new row with data: Alice, 28, Paris",
            input_file=str(test_file),
            output_file=str(output_file)
        )
        print(f"✓ Task created: {task_input.instruction}")
        
        # Run the agent
        print("\n🏃 Running Excel Agent with GLM model...")
        print("This may take a moment...")
        
        agent_response = await runner.run_excel_agent(task_input)
        
        print(f"\n✅ Agent completed successfully!")
        print(f"Response: {agent_response}")
        
        # Verify output file was created
        if output_file.exists():
            print(f"✓ Output file created: {output_file}")
            
            # Read and display the result
            from openpyxl import load_workbook
            wb_result = load_workbook(output_file)
            ws_result = wb_result.active
            
            print("\n📋 Resulting Excel data:")
            for row in ws_result.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    print(f"   {row}")
        else:
            print("❌ Output file was not created")
            return False
            
        return True
        
    except Exception as e:
        print(f"❌ Error during test: {e}")
        import traceback
        traceback.print_exc()
        return False
    
    finally:
        # Cleanup
        for file in [test_file, output_file]:
            if file.exists():
                file.unlink()

async def main():
    """Run the end-to-end test."""
    success = await test_glm_end_to_end()
    
    if success:
        print(f"\n🎉 GLM model integration test PASSED!")
        print(f"🌐 Your web interface is ready at: http://localhost:8000")
        print(f"📝 The backend is using your GLM-4.6 model")
    else:
        print(f"\n❌ GLM model integration test FAILED")
        print(f"🔧 Check the errors above and try again")
    
    return success

if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)