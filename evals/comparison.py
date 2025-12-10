"""
Cell and workbook comparison utilities. Taken from SpreadsheetBench.

Provides functions to compare Excel cell values, styles, and entire workbooks.
"""

import datetime
import os
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill


def datetime_to_float(dt: datetime.datetime) -> float:
    """Convert datetime to Excel serial date format."""
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


# Common date formats to try when parsing date strings
DATE_FORMATS = [
    # European formats (day first)
    "%d.%m.%y",  # 16.02.24
    "%d.%m.%Y",  # 16.02.2024
    "%d/%m/%y",  # 16/02/24
    "%d/%m/%Y",  # 16/02/2024
    "%d-%m-%y",  # 16-02-24
    "%d-%m-%Y",  # 16-02-2024
    # ISO formats
    "%Y-%m-%d",  # 2024-02-16
    "%Y/%m/%d",  # 2024/02/16
    "%Y.%m.%d",  # 2024.02.16
    # US formats (month first)
    "%m/%d/%y",  # 02/16/24
    "%m/%d/%Y",  # 02/16/2024
    "%m-%d-%y",  # 02-16-24
    "%m-%d-%Y",  # 02-16-2024
    # With time components
    "%d.%m.%y %H:%M:%S",
    "%d.%m.%Y %H:%M:%S",
    "%Y-%m-%d %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%m/%d/%Y %H:%M:%S",
]


def try_parse_date_string(s: str) -> datetime.datetime | None:
    """
    Try to parse a string as a date using common formats.

    Returns:
        datetime.datetime if parsing succeeds, None otherwise
    """
    s = s.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def transform_value(v: Any) -> Any:
    """
    Normalize a cell value for comparison.

    - Rounds floats to 2 decimal places
    - Converts datetime.time to string (HH:MM)
    - Converts datetime.datetime to Excel serial date
    - Converts percentage strings (e.g., '99%') to decimal (0.99)
    - Attempts to parse numeric strings
    - Attempts to parse date strings into Excel serial dates
    """
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        # First try to parse as a percentage (e.g., '99%' -> 0.99)
        if v.endswith("%"):
            try:
                v = float(v[:-1]) / 100
            except ValueError:
                pass

        # Then try to parse as a number
        try:
            v = round(float(v), 2)
        except ValueError:
            # Then try to parse as a date
            parsed_date = try_parse_date_string(v)
            if parsed_date is not None:
                v = round(datetime_to_float(parsed_date), 0)
    return v


def compare_cell_values(v1: Any, v2: Any) -> bool:
    """
    Compare two cell values with normalization.

    Handles:
    - Empty string vs None equivalence
    - Numeric rounding
    - Date/time conversion
    - Case-insensitive string comparison
    """
    v1 = transform_value(v1)
    v2 = transform_value(v2)

    # Empty equivalence
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True

    # Type mismatch
    if type(v1) is not type(v2):
        return False

    # Case-insensitive string comparison
    if isinstance(v1, str) and isinstance(v2, str):
        return v1.lower() == v2.lower()

    return v1 == v2


def get_color_rgb(color) -> str:
    """Extract RGB value from color object, defaulting to '00000000' if not valid."""
    if color is None:
        return "00000000"
    
    # Check for theme color (common in Excel files)
    if color.theme is not None:
        return f"theme:{color.theme}"
    
    # Check for indexed color
    if color.indexed is not None:
        return f"indexed:{color.indexed}"
    
    # Check RGB value
    if color.rgb and isinstance(color.rgb, str) and color.rgb != "00000000":
        return color.rgb
    
    return "00000000"


def get_fill_color(fill: PatternFill) -> str:
    """Get the effective fill color, handling solid fills correctly."""
    if fill is None or fill.patternType is None:
        return "none"
    
    if fill.patternType == "solid":
        # For solid fills, the color is in start_color/fgColor
        return get_color_rgb(fill.start_color)
    
    return f"pattern:{fill.patternType}"


def compare_colors(color1, color2) -> bool:
    """Compare two colors using only last 6 characters (RGB), ignoring alpha channel."""
    rgb1 = get_color_rgb(color1)
    rgb2 = get_color_rgb(color2)
    # Handle theme/indexed colors
    if rgb1.startswith("theme:") or rgb1.startswith("indexed:"):
        return rgb1 == rgb2
    if rgb2.startswith("theme:") or rgb2.startswith("indexed:"):
        return rgb1 == rgb2
    return rgb1[-6:] == rgb2[-6:]


def compare_fill_color(fill1: PatternFill, fill2: PatternFill) -> bool:
    """Compare fill colors between two cells."""
    color1 = get_fill_color(fill1)
    color2 = get_fill_color(fill2)
    return color1 == color2


def compare_font_color(font1: Font, font2: Font) -> bool:
    """Compare font colors between two fonts."""
    return compare_colors(font1.color, font2.color)


def col_num_to_name(n: int) -> str:
    """Convert a column number (1-indexed) to an Excel column name."""
    name = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name_to_num(name: str) -> int:
    """Convert an Excel column name to a column number (1-indexed)."""
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord("A") + 1)
    return num


def parse_cell_range(range_str: str) -> tuple[tuple[int, int], tuple[int, int]]:
    """
    Parse a range string like 'A1:AB12'.

    Returns:
        Tuple of ((start_col, start_row), (end_col, end_row))
    """
    start_cell, end_cell = range_str.split(":")

    start_col, start_row = "", ""
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char

    end_col, end_row = "", ""
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name_to_num(start_col), int(start_row)), (
        col_name_to_num(end_col),
        int(end_row),
    )


def generate_cell_names(range_str: str) -> list[str]:
    """Generate a list of all cell names in the specified range."""
    if ":" not in range_str:
        return [range_str]

    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num_to_name(i) for i in range(start_col, end_col + 1)]
    cell_names = [
        f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)
    ]
    return cell_names


class CellComparisonResult:
    """Result of comparing a range of cells."""

    def __init__(self):
        self.is_match = True
        self.correct_count = 0
        self.total_count = 0
        self.differences: list[dict] = []

    @property
    def accuracy(self) -> float:
        """Return the accuracy as a ratio."""
        if self.total_count == 0:
            return 1.0
        return self.correct_count / self.total_count

    def add_match(self):
        """Record a matching cell."""
        self.correct_count += 1
        self.total_count += 1

    def add_difference(
        self, cell_name: str, expected: Any, actual: Any, sheet_name: str = ""
    ):
        """Record a cell difference."""
        self.is_match = False
        self.total_count += 1
        self.differences.append(
            {
                "cell": f"{sheet_name}!{cell_name}" if sheet_name else cell_name,
                "expected": expected,
                "actual": actual,
            }
        )

    def __bool__(self) -> bool:
        return self.is_match

    def __str__(self) -> str:
        if self.is_match:
            return f"Match: {self.correct_count}/{self.total_count} cells"

        diff_msgs = [
            f"  {d['cell']}: expected {d['expected']!r}, got {d['actual']!r}"
            for d in self.differences[:10]
        ]
        if len(self.differences) > 10:
            diff_msgs.append(f"  ... and {len(self.differences) - 10} more")

        return (
            f"Mismatch: {self.correct_count}/{self.total_count} cells correct. \n"
            + "\n".join(diff_msgs)
        )


def get_cf_rule_key(rule) -> str:
    """Generate a key to identify a conditional formatting rule by its condition."""
    parts = [rule.type]
    if hasattr(rule, 'operator') and rule.operator:
        parts.append(rule.operator)
    if hasattr(rule, 'formula') and rule.formula:
        parts.append(str(rule.formula))
    return "|".join(parts)


def get_dxf_fill_color(dxf) -> str:
    """Extract fill color from a DifferentialStyle."""
    if dxf is None or dxf.fill is None:
        return "none"
    fill = dxf.fill
    if fill.patternType == "solid" or fill.fgColor:
        color = fill.fgColor
        if color:
            if color.rgb and color.rgb != "00000000":
                return f"rgb:{color.rgb}"
            if color.theme is not None:
                return f"theme:{color.theme}"
            if color.indexed is not None:
                return f"indexed:{color.indexed}"
    return "none"


def get_dxf_font_style(dxf) -> dict:
    """Extract font style from a DifferentialStyle."""
    if dxf is None or dxf.font is None:
        return {"bold": None, "italic": None, "color": None}
    font = dxf.font
    result = {
        "bold": font.bold,
        "italic": font.italic,
        "color": None,
    }
    if font.color:
        if font.color.rgb and font.color.rgb != "00000000":
            result["color"] = f"rgb:{font.color.rgb}"
        elif font.color.theme is not None:
            result["color"] = f"theme:{font.color.theme}"
    return result


def compare_conditional_formatting(ws_expected, ws_actual, result: CellComparisonResult):
    """
    Compare conditional formatting between worksheets.
    
    Simple check: if expected has CF rules, actual should also have CF rules
    covering similar cells with similar formatting.
    """
    from openpyxl.utils import range_boundaries, get_column_letter
    
    def get_cf_cells_and_fills(ws) -> dict[str, set[str]]:
        """Get {cell: set of fill colors} for all cells with CF."""
        cell_fills = {}
        for cf_range in ws.conditional_formatting:
            # Get the actual range string from the CF object
            range_str = cf_range.sqref if hasattr(cf_range, 'sqref') else str(cf_range.cells)
            
            # Get fill colors from rules
            fills = set()
            for rule in ws.conditional_formatting[cf_range]:
                fill = get_dxf_fill_color(rule.dxf)
                if fill != "none":
                    fills.add(fill)
            
            # Expand range to individual cells
            for part in str(range_str).split():
                try:
                    min_col, min_row, max_col, max_row = range_boundaries(part)
                    for row in range(min_row, max_row + 1):
                        for col in range(min_col, max_col + 1):
                            cell = f"{get_column_letter(col)}{row}"
                            if cell not in cell_fills:
                                cell_fills[cell] = set()
                            cell_fills[cell].update(fills)
                except Exception:
                    continue
            
        return cell_fills
    
    expected_cf = get_cf_cells_and_fills(ws_expected)
    actual_cf = get_cf_cells_and_fills(ws_actual)
    
    # If expected has CF but actual doesn't, that's a fail
    if expected_cf and not actual_cf:
        result.add_difference(
            "conditional_formatting",
            f"{len(expected_cf)} cells have CF",
            "no CF rules",
            ws_expected.title,
        )
        return
    
    # Check each cell that has CF in expected
    for cell in sorted(expected_cf.keys()):
        exp_fills = expected_cf[cell]
        act_fills = actual_cf.get(cell, set())
        
        if not act_fills and exp_fills:
            result.add_difference(
                f"CF {cell}",
                f"fills: {sorted(exp_fills)}",
                "no CF",
                ws_expected.title,
            )
        elif exp_fills != act_fills:
            result.add_difference(
                f"CF {cell} fill",
                sorted(exp_fills),
                sorted(act_fills),
                ws_expected.title,
            )


def compare_cell_range(
    ws_expected,
    ws_actual,
    cell_range: str,
    check_style: bool = False,
) -> CellComparisonResult:
    """
    Compare cells in a range between two worksheets.

    Args:
        ws_expected: Expected worksheet
        ws_actual: Actual worksheet
        cell_range: Range string (e.g., 'A1:B5')
        check_style: Whether to also compare cell styles

    Returns:
        CellComparisonResult with comparison details
    """
    result = CellComparisonResult()
    cell_names = generate_cell_names(cell_range)

    for cell_name in cell_names:
        cell_expected = ws_expected[cell_name]
        cell_actual = ws_actual[cell_name]

        if compare_cell_values(cell_expected.value, cell_actual.value):
            result.add_match()
        else:
            result.add_difference(
                cell_name,
                cell_expected.value,
                cell_actual.value,
                ws_expected.title,
            )

        if check_style:
            expected_fill = get_fill_color(cell_expected.fill)
            actual_fill = get_fill_color(cell_actual.fill)
            
            if expected_fill != actual_fill:
                result.add_difference(
                    f"{cell_name} (fill)",
                    expected_fill,
                    actual_fill,
                    ws_expected.title,
                )
    
    # Compare which cells have CF applied and what colors they would show
    if check_style:
        from openpyxl.utils import range_boundaries, get_column_letter
        
        def get_cf_cells_with_fills(ws) -> dict[str, set[str]]:
            """Get cells that have CF and what fill colors are defined."""
            cell_fills = {}
            for cf in ws.conditional_formatting:
                # Get fill colors from rules
                fills = set()
                for rule in ws.conditional_formatting[cf]:
                    if rule.dxf and rule.dxf.fill:
                        fill = rule.dxf.fill
                        if fill.fgColor:
                            c = fill.fgColor
                            if c.rgb and c.rgb not in ("00000000", "000000"):
                                fills.add(c.rgb[-6:])  # last 6 chars (RGB without alpha)
                            elif c.indexed is not None:
                                fills.add(f"idx{c.indexed}")
                
                # Get cells covered by this CF range
                range_str = str(cf.sqref) if hasattr(cf, 'sqref') else ""
                for part in range_str.split():
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(part)
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                cell = f"{get_column_letter(col)}{row}"
                                if cell not in cell_fills:
                                    cell_fills[cell] = set()
                                cell_fills[cell].update(fills)
                    except:
                        pass
            return cell_fills
        
        exp_cf = get_cf_cells_with_fills(ws_expected)
        act_cf = get_cf_cells_with_fills(ws_actual)
        
        # Check: cells with CF in expected should have CF in actual
        missing_cf_cells = []
        for cell in exp_cf:
            if cell not in act_cf:
                missing_cf_cells.append(cell)
        
        if missing_cf_cells:
            result.add_difference(
                "CF missing on cells",
                f"{len(exp_cf)} cells have CF",
                f"missing CF on: {', '.join(sorted(missing_cf_cells)[:10])}{'...' if len(missing_cf_cells) > 10 else ''}",
                ws_expected.title,
            )
    
    return result


def compare_workbooks(
    expected_path: str,
    actual_path: str,
    answer_positions: str,
    check_style: bool = False,
) -> CellComparisonResult:
    """
    Compare two workbooks at specified cell positions.

    Args:
        expected_path: Path to expected (ground truth) workbook
        actual_path: Path to actual (processed) workbook
        answer_positions: Comma-separated list of ranges (e.g., 'Sheet1!A1:B5,Sheet2!C1')
        check_style: Whether to compare cell styles

    Returns:
        CellComparisonResult with comparison details
    """
    if not os.path.exists(actual_path):
        result = CellComparisonResult()
        result.is_match = False
        result.add_difference("file", expected_path, "File not found")
        return result

    try:
        wb_expected = openpyxl.load_workbook(filename=expected_path, data_only=True)
        wb_actual = openpyxl.load_workbook(filename=actual_path, data_only=True)
    except Exception as e:
        result = CellComparisonResult()
        result.is_match = False
        result.add_difference("load", "success", str(e))
        return result

    combined_result = CellComparisonResult()

    sheet_cell_ranges = answer_positions.split(",")

    for sheet_cell_range in sheet_cell_ranges:
        sheet_cell_range = sheet_cell_range.strip()

        if "!" in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split("!")
            sheet_name = sheet_name.strip("'")
        else:
            sheet_name = wb_expected.sheetnames[0]
            cell_range = sheet_cell_range

        cell_range = cell_range.strip("'")

        # Case-insensitive sheet matching (formulas library uppercases names)
        actual_sheet_name = None
        for name in wb_actual.sheetnames:
            if name.lower() == sheet_name.lower():
                actual_sheet_name = name
                break

        if actual_sheet_name is None:
            combined_result.is_match = False
            combined_result.add_difference(
                f"(missing sheet '{sheet_name}')",
                "sheet exists",
                f"sheet not found in output (has: {wb_actual.sheetnames})",
            )
            continue

        ws_expected = wb_expected[sheet_name]
        ws_actual = wb_actual[actual_sheet_name]

        range_result = compare_cell_range(
            ws_expected, ws_actual, cell_range, check_style
        )

        combined_result.correct_count += range_result.correct_count
        combined_result.total_count += range_result.total_count
        combined_result.differences.extend(range_result.differences)
        if not range_result.is_match:
            combined_result.is_match = False

    return combined_result
