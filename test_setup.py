#!/usr/bin/env python3
"""Test script to verify Excel Agent setup."""

import asyncio
import sys
from pathlib import Path

# Test imports
try:
    from excel_agent.agent_runner import ExcelAgentRunner, TaskInput
    from excel_agent.config import ExperimentConfig
    from excel_mcp.excel_server import mcp
    print("✓ All imports successful")
except ImportError as e:
    print(f"✗ Import error: {e}")
    sys.exit(1)

async def test_mcp_server():
    """Test MCP server startup."""
    print("\n=== Testing MCP Server ===")
    try:
        # Test that we can create the MCP server
        print("✓ MCP server object created successfully")
        
        # Test running server (briefly)
        print("Starting MCP server on http://127.0.0.1:8766...")
        server_task = asyncio.create_task(
            mcp.run_sse_async(host="127.0.0.1", port=8766)
        )
        
        # Wait a moment for server to start
        await asyncio.sleep(2)
        
        # Stop the server
        server_task.cancel()
        try:
            await server_task
        except asyncio.CancelledError:
            pass
            
        print("✓ MCP server started and stopped successfully")
        
    except Exception as e:
        print(f"✗ MCP server test failed: {e}")
        return False
    
    return True

def test_config():
    """Test configuration creation."""
    print("\n=== Testing Configuration ===")
    try:
        # Test with a mock model (we won't actually call it)
        config = ExperimentConfig(model="mock-model")
        print("✓ Configuration created successfully")
        
        # Test prompt formatting
        prompt = config.format_prompt("Test instruction", "test.xlsx")
        print("✓ Prompt formatting works")
        print(f"  Prompt length: {len(prompt)} characters")
        
    except Exception as e:
        print(f"✗ Configuration test failed: {e}")
        return False
    
    return True

def test_file_operations():
    """Test basic file operations."""
    print("\n=== Testing File Operations ===")
    try:
        # Check if test files exist
        input_file = Path("test_input.xlsx")
        output_file = Path("test_output.xlsx")
        
        if not input_file.exists():
            print("✗ Test input file not found")
            return False
            
        if not output_file.exists():
            print("✗ Test output file not found")
            return False
            
        print("✓ Test files exist")
        
        # Test reading Excel file
        import openpyxl
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active
        print(f"✓ Excel file loaded successfully")
        print(f"  Sheet name: {ws.title}")
        print(f"  Data range: A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}")
        
    except Exception as e:
        print(f"✗ File operations test failed: {e}")
        return False
    
    return True

async def main():
    """Run all tests."""
    print("🚀 Starting Excel Agent Setup Tests")
    print("=" * 50)
    
    tests = [
        ("Configuration", test_config),
        ("File Operations", test_file_operations),
        ("MCP Server", test_mcp_server),
    ]
    
    results = []
    for test_name, test_func in tests:
        try:
            if asyncio.iscoroutinefunction(test_func):
                result = await test_func()
            else:
                result = test_func()
            results.append((test_name, result))
        except Exception as e:
            print(f"✗ {test_name} test crashed: {e}")
            results.append((test_name, False))
    
    # Summary
    print("\n" + "=" * 50)
    print("📊 Test Results Summary:")
    print("=" * 50)
    
    passed = 0
    total = len(results)
    
    for test_name, result in results:
        status = "✓ PASS" if result else "✗ FAIL"
        print(f"{test_name:20} {status}")
        if result:
            passed += 1
    
    print(f"\nOverall: {passed}/{total} tests passed")
    
    if passed == total:
        print("\n🎉 All tests passed! The Excel Agent is ready to use.")
        print("\nNext steps:")
        print("1. Set up your API key (OpenAI, Anthropic, etc.)")
        print("2. Run the MCP server: python -m excel_mcp.excel_server")
        print("3. Use the agent runner with your preferred model")
        return True
    else:
        print(f"\n❌ {total - passed} test(s) failed. Please check the errors above.")
        return False

if __name__ == "__main__":
    success = asyncio.run(main())
    sys.exit(0 if success else 1)